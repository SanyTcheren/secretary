#! /usr/bin/env python3
"""Helper for create report and send email to boss."""

from pathlib import Path
from datetime import datetime, timedelta
from dataclasses import dataclass
from calendar import monthrange
import subprocess
import shelve
import logging
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl_image_loader import SheetImageLoader
import ezgmail


@dataclass
class Job():
    """Ежедневная работа."""

    place: str = None
    device: str = None
    detail: str = None

    def __str__(self):
        """Данные по работе, без детального описания."""
        return f"{self.place}, {self.device}"


@dataclass
class Worker():
    """Рабочий."""

    name: str = None
    position: str = None
    start: int = 0
    finish: int = 0
    sign: str = None

    def __str__(self):
        """Данные по рабочему."""
        return f"""
    name: {self.name}
    position: {self.position}
    working from {self.start} to {self.finish}
    sign_path: {self.sign}
    """

class Secretary:
    """My beautiful girl."""

    init_path = './init.data' # файл с исходными данными
    staff_path = './staff.xlsx' # файл с данными по персоналу
    oil_path = './blank/oil.xlsx' # файл с образцом отчета по ДЭС
    work_path = './blank/work.xlsx' # файл с образцом отчета оп ППР
    monthly_path = './blank/monthly.xlsx' # файл с образцом месячного отчета ППР

    def __init__(self):
        """Загружает файл с настройками, при отсутствии файла - создает его."""
        logging.info('Create secretary.')
        if not Path(Secretary.init_path).is_file():
            Secretary.init_data()
        init_file = shelve.open(Secretary.init_path)
        self.boss_email = init_file['boss_email']
        self.my_email = init_file['my_email']
        self.brigade = init_file['brigade']
        init_file.close()
        Secretary.check_gmail(self.my_email)

    def get_staff(self):
        """Чтение данных о персонале из файла."""
        workers = []
        staff_wb = openpyxl.load_workbook(Secretary.staff_path)
        staff_sheet = staff_wb.active
        image_loader = SheetImageLoader(staff_sheet)
        for i in range(6):
            start = staff_sheet[f"B{3+i}"].value
            finish = staff_sheet[f"C{3+i}"].value
            try:
                start = int(start)
                finish = int(finish)
                if start >= finish:
                    logging.warning('Время начало работы превышает время окончания.')
                    raise ValueError()
            except (TypeError, ValueError):
                msg = f"Отсутствуют данные по рабочему №{i+1}"
                logging.warning(msg)
                continue
            name = staff_sheet[f"E{3+i}"].value
            position = staff_sheet[f"D{3+i}"].value
            sign = image_loader.get(f'F{3+i}')
            sign_path = f"signs/sign_{i}.png"
            sign.save(sign_path)
            worker = Worker(name, position, start, finish, sign_path)
            workers.append(worker)
        return workers

    def get_workers(self, day, staff=None):
        """Получение списка рабочих по дню."""
        workers = []
        if staff is None:
            staff = self.get_staff()
        for worker in staff:
            if worker.start <= day <= worker.finish:
                workers.append(worker)
        return workers

    def get_job(self, day):
        """Получение работы за текущий день."""
        jobs_wb = openpyxl.load_workbook(Secretary.monthly_path)
        jobs_sheet = jobs_wb.active
        detail = jobs_sheet.cell(row=(1+3*day), column=4).value
        device = jobs_sheet.cell(row=(1+3*day), column=3).value
        place = jobs_sheet.cell(row=(1+3*day), column=2).value
        job = Job(place, device, detail)
        return job

    def daily_report(self, date_of_report):
        """Создание и отправка суточного рапорта."""
        yesterday = (date_of_report - timedelta(days=1)).date() # День работ
        today = date_of_report.date()

        job = self.get_job(yesterday.day)
        workers = self.get_workers(today.day)

        work_wb = openpyxl.load_workbook(Secretary.work_path)
        work_sheet = work_wb.active
        work_sheet['A3'] = today
        work_sheet['D4'] = today
        work_sheet['B4'] = yesterday
        work_sheet['A7'] = str(job)
        work_sheet['C6'] = job.detail
        for inx, worker in enumerate(workers):
            work_sheet[f"F{6+inx}"] = f"{worker.position} {worker.name}"

        work_report = f'reports/{self.brigade}бр_Суточный_отчет_ППР_{today}.xlsx'
        work_wb.save(work_report)

        oil_wb = openpyxl.load_workbook(Secretary.oil_path)
        oil_sheet = oil_wb.active
        oil_sheet['C12'] = today
        oil_sheet['C19'] = yesterday
        oil_report = f'reports/{self.brigade}бр_Суточный_отчет_ДЭС_{today}.xlsx'
        oil_wb.save(oil_report)

        ezgmail.send(self.boss_email,
                     f"{self.brigade}бр Суточный отчет",
                     '',
                     [work_report, oil_report])
        logging.info('Создан и отправлен суточный отчет.')

    def monthly_report(self, today):
        """
        Создание и отправка месячного отчета ППР.
        Месяц оперделяется по вчерашнему дню.
        """
        months = ['январь','февраль','март','апрель','май','июнь','июль',
                 'август','сентябрь','октябрь','ноябрь','декабрь']
        # определяем месяц вчерашнего дня (1-12)
        month = (today - timedelta(days=1)).month
        year = (today - timedelta(days=1)).year
        # определяем колличество дней в месяце
        days = monthrange(year, month)[1]
        staff = self.get_staff()

        monthly_wb = openpyxl.load_workbook(Secretary.monthly_path)
        monthly_sheet = monthly_wb.active
        monthly_sheet['A1'] = f"Отчет ППР {self.brigade}бр за {months[month-1]} {year}года"
        for i in range(1,32):
            row_day = 1 + 3*i
            if i <= days:
                day = datetime(year,month,i).date()
                monthly_sheet['A'+str(row_day)] = day
# Для того что бы не читать файл на каждой итерации, вызываем с get_workers с
# переменной staff
                workers = self.get_workers(day.day, staff)
                for inx, worker in enumerate(workers):
                    monthly_sheet['E'+str(row_day+inx)] = f"{worker.position} {worker.name}"
                    monthly_sheet.add_image(Image(worker.sign), 'F'+str(row_day+inx))
            else:
                count = 3*(32-i) # Колличество лишних строк в отчете
                monthly_sheet.delete_rows(row_day, count)
                break

        monthly_report = f"reports/{self.brigade}бр_Отчет_ППР_за_{months[month-1]}_{year}года"
        monthly_wb.save(monthly_report+'.xlsx')
# Для преобразования отчета в pdf ,будем использовать libreoffice и запускать
# его как субпроцесс Линукс.
        convert = ('libreoffice --headless --convert-to pdf:calc_pdf_Export' +
                   f' --outdir reports/ {monthly_report}.xlsx')
        subprocess.run(convert, shell=True)

        ezgmail.send(self.boss_email,
                     f"{self.brigade}бр Oтчет ППР за {months[month-1]}_{year}года",
                     '',
                     [monthly_report+".pdf"])
        logging.info('Создан и отправлен месячный отчет ППР.')

    def change_workers(self):
        """Меняем даты работ при переходе на следующий месяц."""
        staff_wb = openpyxl.load_workbook(Secretary.staff_path)
        staff_sheet = staff_wb.active
        for i in range(6):
            start = staff_sheet['B'+str(3+i)].value
            end = staff_sheet['C'+str(3+i)].value
            if start == 1 and end != 31:
                staff_sheet['B'+str(3+i)].value = end + 1
                staff_sheet['C'+str(3+i)].value = 31
            if start != 1 and end == 31:
                staff_sheet['B'+str(3+i)].value = 1
                staff_sheet['C'+str(3+i)].value = start - 1
        staff_wb.save(Secretary.staff_path)
        logging.info('Выполнена смена дат в файле staff.xlsx')

    def work(self):
        """Выполнение основных обязаностей."""
        today = datetime.now()
        self.daily_report(today)
        if today.day == 1:
            self.monthly_report(today)
            self.change_workers()
        logging.info('Work completed.')

    @staticmethod
    def init_data():
        """Создание файла с начальными данными."""
        logging.info('Запись начальных данных')
        init_file = shelve.open(Secretary.init_path)
        init_file['boss_email'] = input('boss email:')
        init_file['my_email'] = input('my email:')
        init_file['brigade'] = input('brigade:')
        init_file.close()

    @staticmethod
    def check_gmail(gmail):
        """Проверка связи с gmail."""
        try:
            ezgmail.init()
            if gmail != ezgmail.EMAIL_ADDRESS:
                logging.error('Несовпадает Ваш email c credentials.json')
                raise Exception
            logging.info('Check_gmail successfully completed.')
        except Exception as ex:
            logging.error('Неудалось связаться с gmail, проверьте токен в credentials.json')
            raise ex


def main():
    """Work."""
    day = datetime.now().date()
    logging.basicConfig(filename=f"./logs/{day}",
                        level=logging.INFO,
                        format="%(asctime)s%(levelname)s[%(name)s] %(message)s")
    secretary = Secretary()
    secretary.work()


if __name__ == '__main__':
    main()
